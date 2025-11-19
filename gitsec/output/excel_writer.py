from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from ..models.finding import DependencyFinding, Finding, SecretFinding


class ExcelReportWriter:
    SEVERITY_ORDER = ["critical", "high", "medium", "low", "unknown"]
    SEVERITY_COLORS = {
        "critical": "8B0000",
        "high": "DC143C",
        "medium": "FF8C00",
        "low": "FFD700",
        "unknown": "808080",
    }

    PIE_CHART_COLORS = {
        "critical": "8B0000",
        "high": "E74C3C",
        "medium": "F39C12",
        "low": "F1C40F",
        "unknown": "95A5A6",
    }

    def __init__(self, output_path: Path):
        self.output_path = output_path
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self.wb = Workbook()
        if self.wb.active:
            self.wb.remove(self.wb.active)

    def add_security_findings(
        self, findings: List[Finding], sheet_name: str = "Security Checks"
    ) -> None:
        if not findings:
            return

        sheet_name = sheet_name[:31]
        ws = self.wb.create_sheet(sheet_name)

        self._add_module_summary(ws, findings, "Security Findings")

        headers = [
            "Check ID",
            "Title",
            "Severity",
            "Category",
            "Resource",
            "Evidence",
            "Description",
            "Risk",
            "Remediation",
            "Reference URL",
            "Notes",
            "Timestamp",
        ]

        header_row = 18
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        sorted_findings = self._sort_findings_by_severity(findings)

        row = header_row + 1
        for finding in sorted_findings:
            data = [
                finding.check_id,
                finding.title,
                finding.severity,
                finding.category,
                finding.resource,
                finding.evidence,
                finding.description,
                finding.risk,
                finding.remediation,
                finding.reference_url,
                finding.notes,
                finding.timestamp,
            ]

            for col_num, value in enumerate(data, 1):
                if value is not None:
                    value = str(value)
                    value = (
                        value.replace("\n", " ").replace("\r", " ").replace("\t", " ")
                    )
                    value = "".join(
                        char if char.isprintable() or char == " " else " "
                        for char in value
                    )
                cell = ws.cell(row=row, column=col_num, value=value)

                if col_num == 3 and value and value.lower() in self.SEVERITY_COLORS:
                    cell.fill = PatternFill(
                        start_color=self.SEVERITY_COLORS[value.lower()],
                        end_color=self.SEVERITY_COLORS[value.lower()],
                        fill_type="solid",
                    )
                    cell.font = Font(color="FFFFFF", bold=True)

            row += 1

        if row > header_row + 1:
            table_ref = f"A{header_row}:{get_column_letter(len(headers))}{row - 1}"
            table = Table(displayName=sheet_name.replace(" ", "_"), ref=table_ref)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style
            ws.add_table(table)

        self._auto_adjust_columns(ws)

    def add_dependency_findings(
        self,
        vulnerabilities: List[DependencyFinding],
        deprecated: Optional[List[dict]] = None,
        unpinned: Optional[List[dict]] = None,
    ) -> None:

        if vulnerabilities:
            ws = self.wb.create_sheet("Dependency Vulnerabilities")
            self._add_module_summary(ws, vulnerabilities, "Dependency Vulnerabilities")

            headers = [
                "Repository",
                "Package",
                "Version",
                "Ecosystem",
                "File Path",
                "Severity",
                "Advisory ID",
                "Title",
                "CVSS Score",
                "URL",
            ]

            header_row = 18
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

            row = header_row + 1
            for finding in vulnerabilities:
                data = [
                    finding.repository,
                    finding.package,
                    finding.version,
                    finding.ecosystem,
                    finding.file_path,
                    finding.severity,
                    finding.advisory_id,
                    finding.title,
                    finding.cvss_score,
                    finding.url,
                ]

                for col_num, value in enumerate(data, 1):
                    if value is not None:
                        value = str(value)
                        value = (
                            value.replace("\n", " ")
                            .replace("\r", " ")
                            .replace("\t", " ")
                        )
                        value = "".join(
                            char if char.isprintable() or char == " " else " "
                            for char in value
                        )
                    cell = ws.cell(row=row, column=col_num, value=value)

                    if col_num == 6 and value and value.lower() in self.SEVERITY_COLORS:
                        cell.fill = PatternFill(
                            start_color=self.SEVERITY_COLORS[value.lower()],
                            end_color=self.SEVERITY_COLORS[value.lower()],
                            fill_type="solid",
                        )
                        cell.font = Font(color="FFFFFF", bold=True)

                row += 1

            if row > header_row + 1:
                table_ref = f"A{header_row}:{get_column_letter(len(headers))}{row - 1}"
                table = Table(displayName="Dependency_Vulnerabilities", ref=table_ref)
                style = TableStyleInfo(
                    name="TableStyleMedium9",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                table.tableStyleInfo = style
                ws.add_table(table)

            self._auto_adjust_columns(ws)

        if deprecated:
            ws = self.wb.create_sheet("Deprecated Packages")
            self._add_simple_header(ws, "Deprecated Packages", len(deprecated))

            headers = list(deprecated[0].keys()) if deprecated else []
            header_row = 5
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="70AD47", end_color="70AD47", fill_type="solid"
                )

            row = header_row + 1
            for item in deprecated:
                for col_num, header in enumerate(headers, 1):
                    value = item.get(header, "")
                    if value:
                        value = str(value)
                        value = (
                            value.replace("\n", " ")
                            .replace("\r", " ")
                            .replace("\t", " ")
                        )
                        value = "".join(
                            char if char.isprintable() or char == " " else " "
                            for char in value
                        )
                    ws.cell(row=row, column=col_num, value=value)
                row += 1

            if row > header_row + 1 and headers:
                table_ref = f"A{header_row}:{get_column_letter(len(headers))}{row - 1}"
                table = Table(displayName="Deprecated_Packages", ref=table_ref)
                style = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                table.tableStyleInfo = style
                ws.add_table(table)

            self._auto_adjust_columns(ws)

        if unpinned:
            ws = self.wb.create_sheet("Unpinned Dependencies")
            self._add_simple_header(ws, "Unpinned Dependencies", len(unpinned))

            headers = list(unpinned[0].keys()) if unpinned else []
            header_row = 5
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="70AD47", end_color="70AD47", fill_type="solid"
                )

            row = header_row + 1
            for item in unpinned:
                for col_num, header in enumerate(headers, 1):
                    value = item.get(header, "")
                    if value:
                        value = str(value)
                        value = (
                            value.replace("\n", " ")
                            .replace("\r", " ")
                            .replace("\t", " ")
                        )
                        value = "".join(
                            char if char.isprintable() or char == " " else " "
                            for char in value
                        )
                    ws.cell(row=row, column=col_num, value=value)
                row += 1

            if row > header_row + 1 and headers:
                table_ref = f"A{header_row}:{get_column_letter(len(headers))}{row - 1}"
                table = Table(displayName="Unpinned_Dependencies", ref=table_ref)
                style = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                table.tableStyleInfo = style
                ws.add_table(table)

            self._auto_adjust_columns(ws)

    def add_secret_findings(self, findings: List[SecretFinding]) -> None:
        if not findings:
            return

        ws = self.wb.create_sheet("Secret Findings")
        self._add_simple_header(ws, "Secret Findings", len(findings))

        headers = [
            "Repository",
            "File Path",
            "Secret Type",
            "Line Number",
            "Secret Hash",
        ]
        header_row = 5
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="C00000", end_color="C00000", fill_type="solid"
            )

        row = header_row + 1
        for finding in findings:
            data = [
                finding.repository,
                finding.file_path,
                finding.secret_type,
                finding.line_number,
                finding.secret_hash,
            ]
            for col_num, value in enumerate(data, 1):
                if value is not None:
                    value = str(value)
                    value = (
                        value.replace("\n", " ").replace("\r", " ").replace("\t", " ")
                    )
                    value = "".join(
                        char if char.isprintable() or char == " " else " "
                        for char in value
                    )
                ws.cell(row=row, column=col_num, value=value)
            row += 1

        if row > header_row + 1:
            table_ref = f"A{header_row}:{get_column_letter(len(headers))}{row - 1}"
            table = Table(displayName="Secret_Findings", ref=table_ref)
            style = TableStyleInfo(
                name="TableStyleMedium6",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style
            ws.add_table(table)

        self._auto_adjust_columns(ws)

    def add_summary_sheet(
        self,
        security_findings: Optional[List[Finding]] = None,
        dependency_findings: Optional[List[DependencyFinding]] = None,
        secret_findings: Optional[List[SecretFinding]] = None,
        deprecated_packages: Optional[List[dict]] = None,
        unpinned_dependencies: Optional[List[dict]] = None,
    ) -> None:
        ws = self.wb.create_sheet("Summary", 0)

        ws.merge_cells("A1:E1")
        title_cell = ws["A1"]
        title_cell.value = "Security Posture Report - Summary"
        title_cell.font = Font(size=18, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(
            start_color="2E75B5", end_color="2E75B5", fill_type="solid"
        )
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        overall_counts = self._aggregate_all_findings(
            security_findings, dependency_findings
        )

        row = 3
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"].value = "Overall Statistics"
        ws[f"A{row}"].font = Font(size=14, bold=True)
        row += 1

        severity_start_row = row
        for severity in self.SEVERITY_ORDER:
            count = overall_counts.get(severity, 0)
            ws[f"A{row}"].value = severity.capitalize()
            ws[f"B{row}"].value = count
            ws[f"A{row}"].font = Font(bold=True)

            if severity in self.SEVERITY_COLORS:
                ws[f"A{row}"].fill = PatternFill(
                    start_color=self.SEVERITY_COLORS[severity],
                    end_color=self.SEVERITY_COLORS[severity],
                    fill_type="solid",
                )
                ws[f"A{row}"].font = Font(bold=True, color="FFFFFF")

            row += 1

        severity_end_row = row - 1
        total = sum(overall_counts.values())
        ws[f"A{row}"].value = "TOTAL"
        ws[f"B{row}"].value = total
        ws[f"A{row}"].font = Font(bold=True, size=12)
        ws[f"B{row}"].font = Font(bold=True, size=12)

        if total > 0:
            self._add_pie_chart(
                ws,
                severity_start_row,
                severity_end_row,
                "D4",
                "Overall Severity Distribution",
            )

        row += 3

        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"].value = "Module Breakdown"
        ws[f"A{row}"].font = Font(size=14, bold=True)
        row += 1

        if security_findings:
            row = self._add_module_summary_to_summary_sheet(
                ws, row, "Security Checks", security_findings, show_total=True
            )

        if dependency_findings:
            row = self._add_module_summary_to_summary_sheet(
                ws,
                row,
                "Dependency Vulnerabilities",
                dependency_findings,
                show_total=True,
            )

        if dependency_findings is not None:
            dep_count = len(deprecated_packages) if deprecated_packages else 0
            ws[f"A{row}"].value = "Deprecated Packages"
            ws[f"B{row}"].value = dep_count
            ws[f"A{row}"].font = Font(bold=True, size=11)
            row += 1

            unpinned_count = len(unpinned_dependencies) if unpinned_dependencies else 0
            ws[f"A{row}"].value = "Unpinned Dependencies"
            ws[f"B{row}"].value = unpinned_count
            ws[f"A{row}"].font = Font(bold=True, size=11)
            row += 2

        if secret_findings:
            ws[f"A{row}"].value = "Secret Scanner"
            ws[f"A{row}"].font = Font(bold=True, size=11)
            row += 1
            ws[f"A{row}"].value = "Total Secrets Found"
            ws[f"B{row}"].value = len(secret_findings)
            ws[f"A{row}"].fill = PatternFill(
                start_color="C00000", end_color="C00000", fill_type="solid"
            )
            ws[f"A{row}"].font = Font(color="FFFFFF", bold=True)
            row += 2

        self._auto_adjust_columns(ws)

    def save(self) -> None:
        self.wb.save(self.output_path)

    def _sort_findings_by_severity(self, findings: List[Finding]) -> List[Finding]:
        severity_rank = {sev: i for i, sev in enumerate(self.SEVERITY_ORDER)}
        return sorted(
            findings,
            key=lambda f: severity_rank.get(
                f.severity.lower() if f.severity else "unknown", 99
            ),
        )

    def _add_module_summary(self, ws, findings: List[Any], title: str) -> None:
        ws.merge_cells("A1:E1")
        title_cell = ws["A1"]
        title_cell.value = title
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(
            start_color="2E75B5", end_color="2E75B5", fill_type="solid"
        )
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        severity_counts = self._count_by_severity(findings)

        row = 3
        severity_start_row = row
        for severity in self.SEVERITY_ORDER:
            count = severity_counts.get(severity, 0)
            ws[f"A{row}"].value = severity.capitalize()
            ws[f"B{row}"].value = count
            ws[f"A{row}"].font = Font(bold=True)

            if severity in self.SEVERITY_COLORS:
                ws[f"A{row}"].fill = PatternFill(
                    start_color=self.SEVERITY_COLORS[severity],
                    end_color=self.SEVERITY_COLORS[severity],
                    fill_type="solid",
                )
                ws[f"A{row}"].font = Font(bold=True, color="FFFFFF")

            row += 1

        severity_end_row = row - 1
        total = sum(severity_counts.values())
        ws[f"A{row}"].value = "Total"
        ws[f"B{row}"].value = total
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].font = Font(bold=True)

        if total > 0:
            self._add_pie_chart(
                ws, severity_start_row, severity_end_row, "D3", f"{title} Distribution"
            )

    def _add_simple_header(self, ws, title: str, count: int) -> None:
        ws.merge_cells("A1:E1")
        title_cell = ws["A1"]
        title_cell.value = title
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(
            start_color="2E75B5", end_color="2E75B5", fill_type="solid"
        )
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        ws["A3"].value = f"Total Items: {count}"
        ws["A3"].font = Font(bold=True)

    def _count_by_severity(self, findings: List[Any]) -> Dict[str, int]:
        counts: Dict[str, int] = defaultdict(int)

        for finding in findings:
            if hasattr(finding, "severity") and finding.severity:
                severity = finding.severity.lower()
                counts[severity] += 1

        return dict(counts)

    def _aggregate_all_findings(
        self,
        security_findings: Optional[List[Finding]],
        dependency_findings: Optional[List[DependencyFinding]],
    ) -> Dict[str, int]:
        counts: Dict[str, int] = defaultdict(int)

        if security_findings:
            for finding in security_findings:
                if finding.severity:
                    counts[finding.severity.lower()] += 1

        if dependency_findings:
            for dep_finding in dependency_findings:
                if dep_finding.severity:
                    counts[dep_finding.severity.lower()] += 1

        return dict(counts)

    def _add_module_summary_to_summary_sheet(
        self,
        ws,
        start_row: int,
        module_name: str,
        findings: List[Any],
        show_total: bool = False,
    ) -> int:
        ws[f"A{start_row}"].value = module_name
        ws[f"A{start_row}"].font = Font(bold=True, size=11)
        start_row += 1

        severity_counts = self._count_by_severity(findings)

        for severity in self.SEVERITY_ORDER:
            count = severity_counts.get(severity, 0)
            if count > 0:
                ws[f"A{start_row}"].value = severity.capitalize()
                ws[f"B{start_row}"].value = count

                if severity in self.SEVERITY_COLORS:
                    ws[f"A{start_row}"].fill = PatternFill(
                        start_color=self.SEVERITY_COLORS[severity],
                        end_color=self.SEVERITY_COLORS[severity],
                        fill_type="solid",
                    )
                    ws[f"A{start_row}"].font = Font(color="FFFFFF", bold=True)

                start_row += 1

        if show_total:
            total = sum(severity_counts.values())
            ws[f"A{start_row}"].value = "Total"
            ws[f"B{start_row}"].value = total
            ws[f"A{start_row}"].font = Font(bold=True)
            ws[f"B{start_row}"].font = Font(bold=True)
            start_row += 1

        start_row += 1
        return start_row

    def _add_pie_chart(
        self, ws, data_start_row: int, data_end_row: int, chart_anchor: str, title: str
    ) -> None:
        from openpyxl.chart.series import DataPoint  # type: ignore[attr-defined]

        chart = PieChart()
        labels = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
        data = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_end_row)

        chart.add_data(data)
        chart.set_categories(labels)

        chart.title = title

        chart.height = 7.5
        chart.width = 13

        if chart.title:
            chart.title.layout = None

        if chart.legend:
            chart.legend.position = "r"  # right side

        if chart.series:
            series = chart.series[0]
            for idx, row_num in enumerate(range(data_start_row, data_end_row + 1)):
                severity_label = ws.cell(row=row_num, column=1).value
                if severity_label:
                    severity_key = severity_label.lower()
                    if severity_key in self.PIE_CHART_COLORS:
                        pt = DataPoint(idx=idx)
                        color = self.PIE_CHART_COLORS[severity_key]
                        pt.graphicalProperties.solidFill = color
                        series.dPt.append(pt)

        ws.add_chart(chart, chart_anchor)

    def _auto_adjust_columns(self, ws) -> None:
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
                length + 2, 50
            )
